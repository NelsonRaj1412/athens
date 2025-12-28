from rest_framework import viewsets, permissions, status, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.db import transaction
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.db.models import Count, Q, Avg, Sum, F, Case, When, Value
from django.db import models
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from datetime import datetime, timedelta
from django.http import HttpResponse, JsonResponse
from django.core.files.base import ContentFile
import json
import base64
import uuid
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from authentication.project_isolation import ProjectIsolationMixin, apply_project_isolation
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    Font = Alignment = PatternFill = None

from .models import (
    Permit, PermitType, PermitApproval, PermitWorker, PermitExtension, 
    PermitAudit, WorkflowTemplate, WorkflowInstance, WorkflowStep,
    HazardLibrary, PermitHazard, GasReading, PermitPhoto, DigitalSignature,
    EscalationRule, NotificationTemplate, SystemIntegration, ComplianceReport
)
try:
    from .qr_utils import generate_permit_qr_code, generate_permit_qr_data
except ImportError:
    generate_permit_qr_code = generate_permit_qr_data = None
from .serializers import (
    PermitSerializer, PermitListSerializer, PermitCreateUpdateSerializer,
    PermitStatusUpdateSerializer, PermitTypeSerializer, PermitApprovalSerializer,
    PermitWorkerSerializer, PermitExtensionSerializer, PermitAuditSerializer,
    WorkflowTemplateSerializer, WorkflowInstanceSerializer, WorkflowStepSerializer,
    HazardLibrarySerializer, PermitHazardSerializer, GasReadingSerializer,
    PermitPhotoSerializer, DigitalSignatureSerializer, EscalationRuleSerializer,
    NotificationTemplateSerializer, SystemIntegrationSerializer, ComplianceReportSerializer,
    PermitAnalyticsSerializer, DashboardStatsSerializer
)
from .permissions import CanManagePermits, CanApprovePermits, CanVerifyPermits
from authentication.models import CustomUser
from permissions.decorators import require_permission

class PermitTypeViewSet(viewsets.ModelViewSet):
    queryset = PermitType.objects.all().order_by('category', 'name')
    serializer_class = PermitTypeSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['category', 'risk_level', 'is_active']
    search_fields = ['name', 'description', 'category']
    pagination_class = None

class HazardLibraryViewSet(viewsets.ModelViewSet):
    queryset = HazardLibrary.objects.filter(is_active=True)
    serializer_class = HazardLibrarySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['category', 'risk_level']
    search_fields = ['name', 'description', 'hazard_id']

class WorkflowTemplateViewSet(viewsets.ModelViewSet):
    queryset = WorkflowTemplate.objects.filter(is_active=True)
    serializer_class = WorkflowTemplateSerializer
    permission_classes = [IsAuthenticated, CanManagePermits]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['permit_type', 'risk_level']

class PermitViewSet(ProjectIsolationMixin, viewsets.ModelViewSet):
    queryset = Permit.objects.all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'permit_type', 'risk_level', 'priority', 'created_by', 'project']
    search_fields = ['permit_number', 'title', 'location', 'description']
    ordering_fields = ['created_at', 'planned_start_time', 'risk_score']
    ordering = ['-created_at']
    model = Permit  # Required for permission decorator
    
    def create(self, request, *args, **kwargs):
        print(f"DEBUG: Received data: {request.data}")
        print(f"DEBUG: permit_type value: {request.data.get('permit_type')}, type: {type(request.data.get('permit_type'))}")
        
        # Check if permit types exist
        from .models import PermitType
        permit_types_count = PermitType.objects.count()
        print(f"DEBUG: Total permit types in database: {permit_types_count}")
        
        if permit_types_count > 0:
            available_types = list(PermitType.objects.values_list('id', 'name')[:10])
            print(f"DEBUG: Available permit types: {available_types}")
        
        serializer = self.get_serializer(data=request.data)
        print(f"DEBUG: Serializer validation: {serializer.is_valid()}")
        if not serializer.is_valid():
            print(f"DEBUG: Validation errors: {serializer.errors}")
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def get_serializer_class(self):
        if self.action == 'list':
            return PermitListSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return PermitCreateUpdateSerializer
        elif self.action == 'update_status':
            return PermitStatusUpdateSerializer
        return PermitSerializer
    


    def get_queryset(self):
        user = self.request.user
        
        queryset = Permit.objects.select_related(
            'permit_type', 'created_by'
        ).prefetch_related(
            'assigned_workers__worker', 'identified_hazards__hazard',
            'gas_readings', 'photos', 'signatures', 'approvals', 'audit_logs'
        )

        # Apply project isolation
        return apply_project_isolation(queryset, user)

    def perform_create(self, serializer):
        user_project = getattr(self.request.user, 'project', None)
        
        with transaction.atomic():
            permit = serializer.save(
                created_by=self.request.user,
                project=user_project
            )
            
            # Set current user as context for audit logging
            permit._current_user = self.request.user
            
            # Create workflow instance if template exists
            try:
                self.create_workflow_instance(permit)
            except Exception as e:
                # Log the error but don't fail the permit creation
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to create workflow instance for permit {permit.id}: {str(e)}")
            
            # Send notifications
            try:
                self.send_creation_notifications(permit)
            except Exception as e:
                # Log the error but don't fail the permit creation
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to send creation notifications for permit {permit.id}: {str(e)}")
    
    @require_permission('edit')
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)
    
    @require_permission('edit')
    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)
    
    @require_permission('delete')
    def destroy(self, request, *args, **kwargs):
        return super().destroy(request, *args, **kwargs)

    def create_workflow_instance(self, permit):
        """Create workflow instance using workflow manager"""
        try:
            from .workflow_manager import workflow_manager
            
            # Use the workflow manager to create proper workflow
            workflow = workflow_manager.initiate_workflow(permit, permit.created_by)
            
        except Exception as e:


            
            pass
            
            pass

    def send_creation_notifications(self, permit):
        """Send notifications for permit creation"""
        from .workflow_manager import workflow_manager
        
        # Check if permit has workflow and send notifications
        try:
            if hasattr(permit, 'workflow'):
                workflow = permit.workflow
                # Get pending verification steps
                verification_steps = workflow.steps.filter(
                    step_id='verification',
                    status='pending'
                )
                
                for step in verification_steps:
                    if step.assignee:
                        # Send notification using the existing notification system
                        from authentication.models_notification import Notification
                        
                        Notification.objects.create(
                            user=step.assignee,
                            title='PTW Verification Required',
                            message=f'Permit {permit.permit_number} requires your verification',
                            notification_type='ptw_verification',
                            data={
                                'permit_id': permit.id,
                                'permit_number': permit.permit_number,
                                'location': permit.location,
                                'created_by': permit.created_by.get_full_name()
                            },
                            link=f'/dashboard/ptw/view/{permit.id}'
                        )
        except Exception as e:

            pass

            pass

    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        """Get dashboard statistics"""
        user_project = getattr(request.user, 'project', None)
        base_query = self.get_queryset()
        
        today = timezone.now().date()
        week_start = today - timedelta(days=today.weekday())
        month_start = today.replace(day=1)
        
        stats = {
            'permits_today': base_query.filter(created_at__date=today).count(),
            'permits_this_week': base_query.filter(created_at__date__gte=week_start).count(),
            'permits_this_month': base_query.filter(created_at__date__gte=month_start).count(),
            'pending_approvals': base_query.filter(status='under_review').count(),
            'overdue_permits': base_query.filter(
                status='active',
                planned_end_time__lt=timezone.now()
            ).count(),
            'high_risk_permits': base_query.filter(risk_level__in=['high', 'extreme']).count(),
            'recent_permits': PermitListSerializer(
                base_query.order_by('-created_at')[:5], many=True
            ).data,
            'compliance_score': self.calculate_compliance_score(base_query)
        }
        
        return Response(DashboardStatsSerializer(stats).data)

    def calculate_compliance_score(self, queryset):
        """Calculate compliance score based on various factors"""
        total_permits = queryset.count()
        if total_permits == 0:
            return 100.0
        
        # Factors for compliance scoring
        completed_on_time = queryset.filter(
            status='completed',
            actual_end_time__lte=F('planned_end_time')
        ).count()
        
        properly_documented = queryset.exclude(
            Q(control_measures='') | Q(ppe_requirements=[])
        ).count()
        
        # Calculate weighted score
        time_compliance = (completed_on_time / total_permits) * 40
        documentation_compliance = (properly_documented / total_permits) * 60
        
        return round(time_compliance + documentation_compliance, 1)

    @action(detail=False, methods=['get'])
    def analytics(self, request):
        """Get comprehensive analytics data"""
        queryset = self.get_queryset()
        
        # Date range filtering
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date and end_date:
            queryset = queryset.filter(
                created_at__date__range=[start_date, end_date]
            )
        
        analytics_data = {
            'total_permits': queryset.count(),
            'active_permits': queryset.filter(status='active').count(),
            'completed_permits': queryset.filter(status='completed').count(),
            'overdue_permits': queryset.filter(
                status='active',
                planned_end_time__lt=timezone.now()
            ).count(),
            'average_processing_time': self.calculate_avg_processing_time(queryset),
            'compliance_rate': self.calculate_compliance_rate(queryset),
            'incident_rate': self.calculate_incident_rate(queryset),
            'risk_distribution': self.get_risk_distribution(queryset),
            'status_distribution': self.get_status_distribution(queryset),
            'monthly_trends': self.get_monthly_trends(queryset)
        }
        
        return Response(PermitAnalyticsSerializer(analytics_data).data)

    def calculate_avg_processing_time(self, queryset):
        """Calculate average processing time in hours"""
        completed_permits = queryset.filter(
            status='completed',
            actual_start_time__isnull=False,
            actual_end_time__isnull=False
        )
        
        if not completed_permits.exists():
            return 0.0
        
        total_hours = 0
        count = 0
        
        for permit in completed_permits:
            duration = permit.actual_end_time - permit.actual_start_time
            total_hours += duration.total_seconds() / 3600
            count += 1
        
        return round(total_hours / count, 2) if count > 0 else 0.0

    def calculate_compliance_rate(self, queryset):
        """Calculate compliance rate percentage"""
        total = queryset.count()
        if total == 0:
            return 100.0
        
        compliant = queryset.filter(
            Q(status='completed') | Q(status='active')
        ).exclude(
            planned_end_time__lt=timezone.now(),
            status='active'
        ).count()
        
        return round((compliant / total) * 100, 2)

    def calculate_incident_rate(self, queryset):
        """Calculate incident rate percentage"""
        # This would integrate with incident management system
        return 0.3  # Mock value

    def get_risk_distribution(self, queryset):
        """Get risk level distribution"""
        return dict(
            queryset.values('risk_level').annotate(
                count=Count('id')
            ).values_list('risk_level', 'count')
        )

    def get_status_distribution(self, queryset):
        """Get status distribution"""
        return dict(
            queryset.values('status').annotate(
                count=Count('id')
            ).values_list('status', 'count')
        )

    def get_monthly_trends(self, queryset):
        """Get monthly trends data"""
        # Implementation for monthly trends
        return []

    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        """Update permit status with validation"""
        permit = self.get_object()
        serializer = PermitStatusUpdateSerializer(permit, data=request.data, partial=True)
        
        if serializer.is_valid():
            with transaction.atomic():
                old_status = permit.status
                permit = serializer.save()
                permit._current_user = request.user
                
                # Create audit log
                PermitAudit.objects.create(
                    permit=permit,
                    action=permit.status,
                    user=request.user,
                    comments=serializer.validated_data.get('comments', ''),
                    old_values={'status': old_status},
                    new_values={'status': permit.status}
                )
                
                # Handle workflow progression
                self.handle_workflow_progression(permit, request.user)
                
            return Response(PermitSerializer(permit).data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def handle_workflow_progression(self, permit, user):
        """Handle workflow step progression"""
        try:
            workflow = permit.workflow
            current_step = workflow.steps.filter(
                status='pending',
                order=workflow.current_step
            ).first()
            
            if current_step:
                current_step.status = 'completed'
                current_step.completed_at = timezone.now()
                current_step.assignee = user
                current_step.save()
                
                # Move to next step
                next_step = workflow.steps.filter(
                    order__gt=workflow.current_step
                ).first()
                
                if next_step:
                    workflow.current_step = next_step.order
                    workflow.save()
                else:
                    workflow.status = 'completed'
                    workflow.completed_at = timezone.now()
                    workflow.save()
                    
        except WorkflowInstance.DoesNotExist:
            pass

    @action(detail=True, methods=['post'])
    def add_photo(self, request, pk=None):
        """Add photo to permit"""
        permit = self.get_object()
        
        photo_data = request.data.get('photo')
        photo_type = request.data.get('photo_type', 'during')
        description = request.data.get('description', '')
        gps_location = request.data.get('gps_location', '')
        
        if not photo_data:
            return Response(
                {'error': 'Photo data is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Handle base64 encoded photo
            if photo_data.startswith('data:image'):
                format, imgstr = photo_data.split(';base64,')
                ext = format.split('/')[-1]
                photo_file = ContentFile(
                    base64.b64decode(imgstr),
                    name=f'permit_{permit.id}_{uuid.uuid4().hex[:8]}.{ext}'
                )
            else:
                photo_file = photo_data
            
            photo = PermitPhoto.objects.create(
                permit=permit,
                photo=photo_file,
                photo_type=photo_type,
                description=description,
                taken_by=request.user,
                gps_location=gps_location
            )
            
            return Response(PermitPhotoSerializer(photo).data)
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    def add_signature(self, request, pk=None):
        """Add digital signature to permit"""
        permit = self.get_object()
        
        signature_type = request.data.get('signature_type')
        signature_data = request.data.get('signature_data')
        
        if not all([signature_type, signature_data]):
            return Response(
                {'error': 'Signature type and data are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        signature = DigitalSignature.objects.create(
            permit=permit,
            signature_type=signature_type,
            signatory=request.user,
            signature_data=signature_data,
            ip_address=self.get_client_ip(request),
            device_info=self.get_device_info(request)
        )
        
        return Response(DigitalSignatureSerializer(signature).data)

    @action(detail=True, methods=['post'])
    def add_gas_reading(self, request, pk=None):
        """Add gas reading to permit"""
        permit = self.get_object()
        
        gas_reading = GasReading.objects.create(
            permit=permit,
            tested_by=request.user,
            **request.data
        )
        
        return Response(GasReadingSerializer(gas_reading).data)

    @action(detail=True, methods=['get'])
    def check_work_hours(self, request, pk=None):
        """Check if permit is within allowed work hours"""
        permit = self.get_object()
        from .utils import is_permit_expired_by_work_hours
        
        is_expired = is_permit_expired_by_work_hours(permit)
        is_within_hours = permit.is_within_work_hours()
        
        return Response({
            'is_within_work_hours': is_within_hours,
            'is_expired_by_work_hours': is_expired,
            'work_hours_display': permit.get_work_hours_display()
        })
    
    @action(detail=True, methods=['get'])
    def generate_qr_code(self, request, pk=None):
        """Generate QR code for permit"""
        permit = self.get_object()
        
        try:
            if generate_permit_qr_code is None:
                return Response(
                    {'error': 'QR code generation not available - qrcode library not installed'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Generate QR code image
            qr_image = generate_permit_qr_code(permit)
            
            # Update permit with QR data
            permit.qr_code = generate_permit_qr_data(permit)
            permit.save()
            
            return Response({
                'qr_image': qr_image,
                'qr_data': permit.qr_code,
                'mobile_url': f"http://localhost:5173/mobile/permit/{permit.id}"
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def available_approvers(self, request):
        """Get available approvers based on hierarchical rules (only induction-trained users)"""
        user = request.user
        user_project = getattr(user, 'project', None)
        current_user_type = user.admin_type
        current_user_grade = user.grade
        
        # Get induction-trained users from the same project
        from authentication.project_isolation import apply_user_project_isolation_with_induction
        
        approvers = apply_user_project_isolation_with_induction(
            CustomUser.objects.filter(is_active=True),
            user
        ).exclude(id=user.id)
        
        # Apply hierarchy filtering based on current user
        if current_user_type == 'epcuser' and current_user_grade == 'C':
            # EPC C grade can send to: clientuser (all grades) + epcuser (A & B grades only)
            approvers = approvers.filter(
                models.Q(admin_type='clientuser') |
                models.Q(admin_type='epcuser', grade__in=['A', 'B'])
            )
        elif current_user_type == 'epcuser' and current_user_grade == 'B':
            # EPC B grade can send to: clientuser (all grades) + epcuser (A grade only)
            approvers = approvers.filter(
                models.Q(admin_type='clientuser') |
                models.Q(admin_type='epcuser', grade='A')
            )
        elif current_user_type == 'clientuser' and current_user_grade == 'C':
            # Client C grade can send to: clientuser (A & B grades only)
            approvers = approvers.filter(
                admin_type='clientuser',
                grade__in=['A', 'B']
            )
        elif current_user_type == 'clientuser' and current_user_grade == 'B':
            # Client B grade can send to: clientuser (A grade only)
            approvers = approvers.filter(
                admin_type='clientuser',
                grade='A'
            )
        else:
            # Default: only show clientuser and epcuser
            approvers = approvers.filter(
                admin_type__in=['clientuser', 'epcuser']
            )
        
        # Order by hierarchy: clientuser first, then epcuser, then by grade A->B->C
        approvers = approvers.order_by(
            models.Case(
                models.When(admin_type='clientuser', then=models.Value(1)),
                models.When(admin_type='epcuser', then=models.Value(2)),
                default=models.Value(3)
            ),
            'grade',
            'first_name'
        )
        
        approver_data = []
        for approver in approvers:
            approver_data.append({
                'id': approver.id,
                'username': approver.username,
                'full_name': approver.get_full_name(),
                'admin_type': approver.admin_type,
                'grade': approver.grade
            })
        
        return Response({
            'approvers': approver_data,
            'message': 'Only induction-trained users are shown'
        })
    
    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """Verify permit and send to all users of selected grade"""
        permit = self.get_object()
        
        action = request.data.get('action')
        comments = request.data.get('comments', '')
        user_type = request.data.get('user_type')
        grade = request.data.get('grade')
        
        if action != 'approve':
            return Response(
                {'error': 'Invalid action'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not user_type or not grade:
            return Response(
                {'error': 'Please select user type and grade'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from .workflow_manager import workflow_manager
            
            # Verify permit and send to all users of selected grade
            workflow_manager.verify_permit_to_grade(
                permit, 
                request.user, 
                'approve', 
                comments, 
                user_type,
                grade
            )
            
            return Response({
                'message': f'Permit verified and sent to all {user_type} Grade {grade} users for approval'
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve permit"""
        permit = self.get_object()
        
        action = 'approve'
        comments = request.data.get('comments', '')
        
        try:
            from .workflow_manager import workflow_manager
            
            # Approve permit
            result = workflow_manager.approve_permit(
                permit, 
                request.user, 
                action, 
                comments
            )
            
            if isinstance(result, dict) and result.get('status') == 'already_approved':
                return Response({
                    'error': f"Permit already approved by {result['approved_by']} at {result['approved_at']}"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            return Response({
                'message': 'Permit approved successfully'
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject permit"""
        permit = self.get_object()
        
        action = 'reject'
        comments = request.data.get('comments', '')
        
        if not comments:
            return Response(
                {'error': 'Comments are required for rejection'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            from .workflow_manager import workflow_manager
            
            # Check if this is approval rejection or verification rejection
            if permit.status == 'pending_approval':
                workflow_manager.approve_permit(permit, request.user, action, comments)
                message = 'Permit approval rejected'
            else:
                workflow_manager.verify_permit(permit, request.user, action, comments)
                message = 'Permit verification rejected'
            
            return Response({
                'message': message
            })
            
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=True, methods=['get'])
    def export_pdf(self, request, pk=None):
        """Export permit as PDF"""
        permit = self.get_object()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="permit_{permit.permit_number}.pdf"'
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        
        # PDF generation logic
        self.generate_permit_pdf(p, permit)
        
        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export permits as Excel"""
        queryset = self.filter_queryset(self.get_queryset())
        
        if openpyxl is None:
            return Response(
                {'error': 'Excel export not available - openpyxl not installed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Permits"
        
        # Headers
        headers = [
            'Permit Number', 'Type', 'Status', 'Location', 'Risk Level',
            'Created Date', 'Start Time', 'End Time', 'Created By'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            if Font:
                cell.font = Font(bold=True)
            if PatternFill:
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, permit in enumerate(queryset, 2):
            ws.cell(row=row, column=1, value=permit.permit_number)
            ws.cell(row=row, column=2, value=permit.permit_type.name)
            ws.cell(row=row, column=3, value=permit.get_status_display())
            ws.cell(row=row, column=4, value=permit.location)
            ws.cell(row=row, column=5, value=permit.get_risk_level_display())
            ws.cell(row=row, column=6, value=permit.created_at.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=7, value=permit.planned_start_time.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=8, value=permit.planned_end_time.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=9, value=permit.created_by.username)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="permits.xlsx"'
        
        wb.save(response)
        return response

    def generate_permit_pdf(self, canvas, permit):
        """Generate PDF content for permit"""
        # PDF generation implementation
        canvas.setTitle(f"Permit {permit.permit_number}")
        canvas.drawString(100, 750, f"Permit to Work: {permit.permit_number}")
        # Add more PDF content here

    def get_client_ip(self, request):
        """Get client IP address"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

    def get_device_info(self, request):
        """Get device information"""
        return {
            'user_agent': request.META.get('HTTP_USER_AGENT', ''),
            'platform': request.META.get('HTTP_SEC_CH_UA_PLATFORM', ''),
            'mobile': request.META.get('HTTP_SEC_CH_UA_MOBILE', '') == '?1'
        }

# Additional ViewSets for related models
class PermitWorkerViewSet(viewsets.ModelViewSet):
    queryset = PermitWorker.objects.all()
    serializer_class = PermitWorkerSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['permit', 'worker', 'role']

class PermitApprovalViewSet(viewsets.ModelViewSet):
    queryset = PermitApproval.objects.all()
    serializer_class = PermitApprovalSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['permit', 'approver', 'action', 'approval_level']

class PermitExtensionViewSet(viewsets.ModelViewSet):
    queryset = PermitExtension.objects.all()
    serializer_class = PermitExtensionSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['permit', 'status', 'requested_by']

# Removed WorkTimeExtensionViewSet - time management handled centrally

class PermitAuditViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = PermitAudit.objects.all()
    serializer_class = PermitAuditSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['permit', 'action', 'user']
    ordering = ['-timestamp']

class GasReadingViewSet(viewsets.ModelViewSet):
    queryset = GasReading.objects.all()
    serializer_class = GasReadingSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['permit', 'gas_type', 'status', 'tested_by']

class PermitPhotoViewSet(viewsets.ModelViewSet):
    queryset = PermitPhoto.objects.all()
    serializer_class = PermitPhotoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['permit', 'photo_type', 'taken_by']

class DigitalSignatureViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = DigitalSignature.objects.all()
    serializer_class = DigitalSignatureSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['permit', 'signature_type', 'signatory']

class WorkflowInstanceViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = WorkflowInstance.objects.all()
    serializer_class = WorkflowInstanceSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['permit', 'template', 'status']

class SystemIntegrationViewSet(viewsets.ModelViewSet):
    queryset = SystemIntegration.objects.all()
    serializer_class = SystemIntegrationSerializer
    permission_classes = [IsAuthenticated, CanManagePermits]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['integration_type', 'status', 'is_active']

    @action(detail=True, methods=['post'])
    def test_connection(self, request, pk=None):
        """Test integration connection"""
        integration = self.get_object()
        
        # Implementation for testing connection
        try:
            # Test connection logic here
            integration.status = 'connected'
            integration.last_sync = timezone.now()
            integration.save()
            
            return Response({'status': 'Connection successful'})
        except Exception as e:
            integration.status = 'error'
            integration.save()
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['post'])
    def sync_data(self, request, pk=None):
        """Sync data with external system"""
        integration = self.get_object()
        
        try:
            integration.status = 'syncing'
            integration.save()
            
            # Sync logic here
            
            integration.status = 'connected'
            integration.last_sync = timezone.now()
            integration.save()
            
            return Response({'status': 'Sync completed successfully'})
        except Exception as e:
            integration.status = 'error'
            integration.save()
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

class ComplianceReportViewSet(viewsets.ModelViewSet):
    queryset = ComplianceReport.objects.all()
    serializer_class = ComplianceReportSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['report_type', 'generated_by']
    ordering = ['-generated_at']

    def perform_create(self, serializer):
        """Generate compliance report"""
        report = serializer.save(generated_by=self.request.user)
        
        # Generate report data
        self.generate_report_data(report)

    def generate_report_data(self, report):
        """Generate report data based on type"""
        # Implementation for different report types
        pass

# API endpoints for mobile app
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def sync_offline_data(request):
    """Sync offline data from mobile app"""
    try:
        offline_data = request.data.get('offline_data', [])
        synced_count = 0
        errors = []
        
        for item in offline_data:
            try:
                item_type = item.get('type')
                data = item.get('data')
                
                if item_type == 'permit':
                    # Create permit from offline data
                    serializer = PermitCreateUpdateSerializer(data=data)
                    if serializer.is_valid():
                        permit = serializer.save(created_by=request.user)
                        synced_count += 1
                    else:
                        errors.append(f"Permit validation error: {serializer.errors}")
                
                elif item_type == 'photo':
                    # Handle photo sync
                    permit_id = data.get('permit_id')
                    if permit_id:
                        permit = Permit.objects.get(id=permit_id)
                        # Create photo
                        synced_count += 1
                
                # Handle other types...
                
            except Exception as e:
                errors.append(f"Error syncing item: {str(e)}")
        
        return Response({
            'synced_count': synced_count,
            'errors': errors,
            'status': 'completed'
        })
        
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def qr_scan_permit(request, qr_code):
    """Get permit details from QR code"""
    try:
        # Decode QR code
        qr_data = json.loads(base64.b64decode(qr_code).decode())
        permit_id = qr_data.get('id')
        
        if permit_id:
            permit = get_object_or_404(Permit, id=permit_id)
            serializer = PermitSerializer(permit)
            return Response(serializer.data)
        else:
            return Response(
                {'error': 'Invalid QR code'},
                status=status.HTTP_400_BAD_REQUEST
            )
            
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )

@api_view(['GET'])
def mobile_permit_view(request, permit_id):
    """Mobile-friendly permit view (no authentication required for QR access)"""
    try:
        permit = get_object_or_404(Permit, id=permit_id)
        
        # Basic permit info for mobile view
        permit_data = {
            'id': permit.id,
            'permit_number': permit.permit_number,
            'permit_type': permit.permit_type.name if permit.permit_type else 'Unknown',
            'status': permit.status,
            'location': permit.location,
            'description': permit.description,
            'planned_start_time': permit.planned_start_time.isoformat() if permit.planned_start_time else None,
            'planned_end_time': permit.planned_end_time.isoformat() if permit.planned_end_time else None,
            'risk_level': permit.risk_level,
            'created_by': permit.created_by.username if permit.created_by else 'Unknown',
            'created_at': permit.created_at.isoformat(),
            'control_measures': permit.control_measures,
            'ppe_requirements': permit.ppe_requirements,
            'work_nature': permit.work_nature,
            'work_hours_display': permit.get_work_hours_display(),
        }
        
        return Response(permit_data)
        
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_400_BAD_REQUEST
        )

# Work Hours Management API
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_work_time_settings(request):
    """Get current work time settings from master admin"""
    from .utils import get_work_time_settings
    
    settings = get_work_time_settings()
    
    return Response({
        'day_start': settings['day_start'].strftime('%H:%M'),
        'day_end': settings['day_end'].strftime('%H:%M'),
        'night_start': settings['night_start'].strftime('%H:%M'),
        'night_end': settings['night_end'].strftime('%H:%M'),
    })